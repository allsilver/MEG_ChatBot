import os
import re
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def save_error_log(error_folder, step_name, failed_files):
    if not failed_files:
        return
    os.makedirs(error_folder, exist_ok=True)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    error_path = os.path.join(error_folder, f"error_log_{step_name}_{timestamp}.txt")
    with open(error_path, 'w', encoding='utf-8') as f:
        f.write(f"[{step_name}] 에러 발생 파일 목록 - {timestamp}\n")
        f.write("=" * 50 + "\n")
        f.write("\n".join(failed_files))
    print(f"에러 로그 저장: {error_path}")


def build_title(file_path: str, raw_data_folder: str) -> str:
    rel_path  = os.path.relpath(file_path, raw_data_folder)
    parts     = rel_path.replace('\\', '/').split('/')
    parts[-1] = os.path.splitext(parts[-1])[0]
    return " > ".join(p.strip() for p in parts if p.strip())


# ── No 컬럼 판별 ───────────────────────────────────────────────────────────────
# 유효한 No: 알파벳 포함, 5자 이내, 공백·하이픈·숫자 조합 허용
# 예: A, B1, C-1, C- 2, B -3, D, H
_NO_RE = re.compile(r'^[A-Za-z][A-Za-z0-9\s\-]*$')

def is_valid_no(val: str) -> bool:
    val = str(val).strip()
    if not val or val.lower() in ('nan', 'none', ''):
        return False
    if len(val) > 5:
        return False
    compact = val.replace(' ', '')
    return bool(_NO_RE.match(compact)) and bool(re.search(r'[A-Za-z]', compact))

def is_skip_no(val: str) -> bool:
    val = str(val).strip()
    if not val or val.lower() in ('nan', 'none', ''):
        return True
    if len(val) > 5:
        return True
    if re.match(r'^\d+$', val):
        return True
    return False


# ── 주석 번호 제거 ─────────────────────────────────────────────────────────────
def remove_footnote_numbers(text: str) -> tuple[str, list[str]]:
    removed = []
    PAT_BOTH        = re.compile(r'\s*\(([1-9][0-9]*)\)')
    PAT_RIGHT_FRONT = re.compile(r'(?:^|(?<=,\s)|(?<=,))([1-9][0-9]*)\)\s*')

    def replace_both(m):
        removed.append(f"({m.group(1)})")
        return ""

    def replace_right(m):
        removed.append(f"{m.group(1)})")
        return ""

    text = PAT_BOTH.sub(replace_both, text)
    text = PAT_RIGHT_FRONT.sub(replace_right, text)
    text = re.sub(r'\s+', ' ', text).strip().strip(', ').strip()
    return text, removed


# ── xlsx 단일 파일 추출 ────────────────────────────────────────────────────────
def extract_from_xlsx(file_path: str, title: str) -> tuple[list[dict], list[dict], str | None, bool]:
    """
    xlsx 파일의 첫 번째 시트에서 체크리스트 데이터를 추출합니다.

    [Guide 조합 규칙]
    케이스1 — 서브헤더 있음 (헤더 바로 다음 행에서 No=빈값, Guide 범위에 값 있음)
        값이 있는 열만 "서브헤더 : 값" 으로 만들고 ", " 로 연결
        단, 값이 하나뿐이면 서브헤더 없이 값만 저장

    케이스2 — 서브헤더 없음, Guide 열이 정확히 2개
        두 열 모두 값 있음 → "열1 : 열2"
        한 열만 값 있음   → 해당 값만
        둘 다 없음        → 빈값

    케이스3 — 서브헤더 없음, Guide 열이 1개 또는 3개 이상
        기존 방식 그대로 ", " 로 연결

    [No 컬럼 규칙]
    - 유효한 No(알파벳 기반, 5자 이내): 행 저장
    - 순수 숫자: 유효 처리 안 함, 에러 플래그 세움
    - 공백·문장: 건너뜀. 유효 No 이후 연속 3회 → 표 종료
    """
    extracted      = []
    review         = []
    TABLE_END_STREAK = 3
    has_numeric_no = False

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], [], "빈 시트", False

        # ── 헤더 탐지 ──────────────────────────────────────────────────
        header_idx = no_col = item_start_col = guide_start_col = guide_end_col = None
        for i, row in enumerate(rows):
            row_vals = [str(v).strip().upper() if v is not None else '' for v in row]
            if 'NO' in row_vals and 'ITEM' in row_vals and 'GUIDE' in row_vals:
                header_idx = i
                for j, val in enumerate(row_vals):
                    if val == 'NO'    and no_col          is None: no_col          = j
                    if val == 'ITEM'  and item_start_col  is None: item_start_col  = j
                    if val == 'GUIDE' and guide_start_col is None: guide_start_col = j
                    if guide_start_col is not None and 'FIGURE' in val:
                        guide_end_col = j
                        break
                if guide_end_col is None:
                    guide_end_col = len(row_vals)
                break

        if header_idx is None:
            return [], [], "헤더(NO/ITEM/GUIDE) 없음", False

        guide_col_count = guide_end_col - guide_start_col

        # ── 서브헤더 탐지 ───────────────────────────────────────────────
        # 헤더 바로 다음 행에서 No=빈값이고 Guide 범위에 값이 있으면 서브헤더 행
        guide_subheaders  = []   # Guide 열별 서브헤더 문자열 (없으면 None)
        first_data_offset = 1    # 서브헤더 없으면 헤더+1, 있으면 헤더+2

        if header_idx + 1 < len(rows):
            peek = rows[header_idx + 1]
            no_empty = (peek[no_col] is None or str(peek[no_col]).strip() == '')
            guide_peek_vals = [
                str(peek[g]).strip() if g < len(peek) and peek[g] is not None else ''
                for g in range(guide_start_col, guide_end_col)
            ]
            if no_empty and any(guide_peek_vals):
                guide_subheaders  = [v if v else None for v in guide_peek_vals]
                first_data_offset = 2

        # ── 데이터 행 순회 ──────────────────────────────────────────────
        data_rows      = rows[header_idx + first_data_offset:]
        item_fill_vals = [None] * (guide_start_col - item_start_col)
        started        = False
        skip_streak    = 0
        current_no     = None   # No ffill 용 (No=None 연속 행에서 이전 No 유지)

        for row in data_rows:
            no_raw = str(row[no_col]).strip() if row[no_col] is not None else ''

            # 순수 숫자 No → 플래그만 세우고 건너뜀
            if re.match(r'^\d+$', no_raw) and no_raw:
                has_numeric_no = True
                if started:
                    skip_streak += 1
                    if skip_streak >= TABLE_END_STREAK:
                        break
                continue

            # No가 유효한 알파벳 기반 값인 경우
            if is_valid_no(no_raw):
                started     = True
                skip_streak = 0
                current_no  = no_raw.replace(' ', '').upper()

                # Item ffill (유효한 값만 업데이트)
                for idx in range(guide_start_col - item_start_col):
                    v = row[item_start_col + idx]
                    if v is not None and str(v).strip().lower() not in ('', 'none', 'nan'):
                        item_fill_vals[idx] = str(v).strip()

            # No가 비어있지만 Guide에 값이 있는 경우 → 이전 No/Item 유지하며 추출
            elif not no_raw and started and current_no is not None:
                # Guide 범위에 실제 값이 있는지 확인
                has_guide_val = any(
                    row[g] is not None and str(row[g]).strip().lower() not in ('', 'none', 'nan')
                    for g in range(guide_start_col, guide_end_col)
                )
                if not has_guide_val:
                    # Guide도 비어있으면 건너뜀 처리
                    skip_streak += 1
                    if skip_streak >= TABLE_END_STREAK:
                        break
                    continue
                # Guide 값이 있으면 skip_streak 리셋하고 계속 처리
                skip_streak = 0
                # current_no, item_fill_vals 는 그대로 유지 (ffill)

            else:
                # No가 문장이거나 처리 불가 → 건너뜀
                if not started:
                    continue
                skip_streak += 1
                if skip_streak >= TABLE_END_STREAK:
                    break
                continue

            # ── 공통 처리: Guide 수집 및 조합 ─────────────────────────
            no_clean  = current_no
            full_item = ", ".join(v for v in item_fill_vals if v)

            guide_raw = [
                str(row[g]).strip()
                if g < len(row) and row[g] is not None
                   and str(row[g]).strip().lower() not in ('', 'none', 'nan')
                else None
                for g in range(guide_start_col, guide_end_col)
            ]
            # 실제 값이 있는 것만 추림
            guide_vals = [v for v in guide_raw if v is not None]

            # ── Guide 조합 ─────────────────────────────────────────────
            # 케이스1: 서브헤더 있음
            #   값이 있는 열만 "서브헤더 : 값" 으로 만들고 ", " 로 연결
            #   값이 하나뿐이면 서브헤더 없이 값만
            if guide_subheaders:
                parts = [
                    (sub, val)
                    for sub, val in zip(guide_subheaders, guide_raw)
                    if val is not None
                ]
                if len(parts) == 0:
                    full_guide = ""
                elif len(parts) == 1:
                    full_guide = parts[0][1]
                else:
                    full_guide = ", ".join(
                        f"{sub} : {val}" if sub else val
                        for sub, val in parts
                    )

            # 케이스2·3: 서브헤더 없음 — 실제 값 개수로 판단
            #   값이 2개 → " : " 로 연결
            #   값이 1개 또는 3개 이상 → ", " 로 연결
            else:
                if len(guide_vals) == 2:
                    full_guide = f"{guide_vals[0]} : {guide_vals[1]}"
                else:
                    full_guide = ", ".join(guide_vals)

            clean_item,  removed_item  = remove_footnote_numbers(full_item)
            clean_guide, removed_guide = remove_footnote_numbers(full_guide) \
                if full_guide else ("", [])

            extracted.append({
                "No":     no_clean,
                "Title":  title,
                "Item":   clean_item,
                "Guide":  clean_guide if clean_guide else None,
                "Reason": ""
            })

            if removed_item or removed_guide:
                review.append({
                    "Title":             title,
                    "No":                no_clean,
                    "원본 Item":         full_item,
                    "정제 Item":         clean_item,
                    "제거된 Item 주석":  ", ".join(removed_item),
                    "원본 Guide":        full_guide,
                    "정제 Guide":        clean_guide,
                    "제거된 Guide 주석": ", ".join(removed_guide),
                })

    except Exception as e:
        return [], [], str(e), False

    return extracted, review, None, has_numeric_no


# ── 전체 처리 ──────────────────────────────────────────────────────────────────
def process_and_save_checklists(domain_data_root, db_key, _csv_folder=None):
    """
    raw_data/<db_key>/ 하위의 모든 xlsx 파일을 직접 읽어
    preprocessed_data_final.xlsx 로 저장합니다.

    _csv_folder: 하위 호환용, 미사용
    """
    raw_data_folder = os.path.join(domain_data_root, 'raw_data', db_key)
    error_folder    = os.path.join(domain_data_root, 'error')
    os.makedirs(error_folder, exist_ok=True)

    all_xlsx_files = []
    for root, dirs, files in os.walk(raw_data_folder):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                all_xlsx_files.append(os.path.join(root, file))

    print(f"총 {len(all_xlsx_files)}개의 xlsx 파일 발견")

    all_extracted_data = []
    review_data        = []
    failed_files       = []
    numeric_no_files   = []   # 순수 숫자 No 가 발견된 파일 목록

    for file_path in all_xlsx_files:
        title = build_title(file_path, raw_data_folder)
        extracted, review, err, has_numeric_no = extract_from_xlsx(file_path, title)
        if err:
            print(f"  스킵 [{err}]: {os.path.basename(file_path)}")
            failed_files.append(f"[{err}] {file_path}")
            continue
        if has_numeric_no:
            numeric_no_files.append(file_path)
        all_extracted_data.extend(extracted)
        review_data.extend(review)

    save_error_log(error_folder, "extract_checklists", failed_files)

    # 순수 숫자 No 발견 파일은 별도 에러 로그로 저장
    if numeric_no_files:
        save_error_log(error_folder, "numeric_no_detected", [
            f"[숫자No감지] {f}" for f in numeric_no_files
        ])
        print(f"  ⚠️  숫자 No 감지 파일: {len(numeric_no_files)}개 → error 폴더 로그 확인")

    if not all_extracted_data:
        print("추출된 데이터가 없습니다.")
        return None

    result_df     = pd.DataFrame(all_extracted_data)
    result_folder = os.path.join(domain_data_root, 'result', db_key)
    os.makedirs(result_folder, exist_ok=True)

    output_path = os.path.join(result_folder, 'preprocessed_data_final.xlsx')
    result_df[["No", "Title", "Item", "Guide", "Reason"]].to_excel(
        output_path, index=False, engine='openpyxl'
    )
    print(f"전처리 완료: {output_path} (총 {len(result_df)}행)")

    if review_data:
        review_df   = pd.DataFrame(review_data)
        review_path = os.path.join(result_folder, 'footnote_review.xlsx')
        review_df.to_excel(review_path, index=False, engine='openpyxl')
        print(f"  주석 제거 검토 파일: {review_path} (총 {len(review_df)}행)")
    else:
        print(f"  주석 제거 검토 대상 없음 ✅")

    return output_path


def run_preprocess(domain_data_root: str, db_key: str) -> bool:
    print(f"\n=== [{db_key}] 전처리 시작 ===")
    result = process_and_save_checklists(domain_data_root, db_key)
    if not result:
        print(f"❌ [{db_key}] 전처리 실패")
        return False
    print(f"✅ [{db_key}] 전처리 완료 → {result}")
    return True


if __name__ == "__main__":
    import sys

    current_dir  = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    data_root    = os.path.join(project_root, 'data')

    for reg_path in [
        os.path.join(current_dir,  'domain_registry.json'),
        os.path.join(project_root, 'domain_registry.json'),
    ]:
        if os.path.exists(reg_path):
            with open(reg_path, encoding='utf-8') as f:
                domain_registry = json.load(f)
            break
    else:
        print("❌ domain_registry.json 을 찾을 수 없습니다.")
        sys.exit(1)

    DOMAIN_KEY = "MEG_STANDARD"
    if DOMAIN_KEY not in domain_registry:
        print(f"❌ domain_registry.json 에 '{DOMAIN_KEY}' 항목이 없습니다.")
        sys.exit(1)

    domain_data_root = os.path.join(data_root, DOMAIN_KEY)

    # db_registry 직접 읽기 (domain_registry의 db_keys 미사용)
    db_registry_path = os.path.join(data_root, DOMAIN_KEY, f'db_registry_{DOMAIN_KEY}.json')
    if not os.path.exists(db_registry_path):
        print(f"❌ db_registry 파일을 찾을 수 없습니다: {db_registry_path}")
        sys.exit(1)
    with open(db_registry_path, encoding='utf-8') as f:
        db_registry = json.load(f)
    available_dbs = list(db_registry.keys())

    print("=" * 50)
    print(f"  MEG_STANDARD 전처리 (preprocess_meg.py)")
    print(f"  사용 가능한 DB: {available_dbs}")
    print("=" * 50)

    keys_input = input(
        "\n처리할 DB 키를 입력하세요."
        "\n  단일 입력, 쉼표로 복수 입력, 또는 all (전체)"
        "\n입력 > "
    ).strip()

    if keys_input.lower() == 'all':
        target_keys = available_dbs
    else:
        target_keys = [k.strip() for k in keys_input.split(',') if k.strip() in available_dbs]
        invalid     = [k.strip() for k in keys_input.split(',') if k.strip() not in available_dbs]
        if invalid:
            print(f"⚠️  사용 가능한 DB 목록에 없는 키 제외: {invalid}")

    if not target_keys:
        print("❌ 유효한 DB 키가 없습니다. 종료합니다.")
        sys.exit(1)

    print(f"\n처리 대상 ({len(target_keys)}개): {target_keys}")
    print("설정 완료. 이후 자동으로 실행됩니다.")
    print("=" * 50)

    success_list, fail_list = [], []
    for db_key in target_keys:
        ok = run_preprocess(domain_data_root, db_key)
        (success_list if ok else fail_list).append(db_key)

    print(f"\n{'='*50}")
    print("전처리 완료 요약")
    print(f"{'='*50}")
    print(f"✅ 성공 ({len(success_list)}개): {success_list}")
    if fail_list:
        print(f"❌ 실패 ({len(fail_list)}개): {fail_list}")
    print(f"\n다음 단계: python src/table_parser.py 실행")
