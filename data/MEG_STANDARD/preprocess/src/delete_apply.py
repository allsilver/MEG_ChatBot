"""
delete_apply.py
---------------
name_analysis.xlsx 의 삭제 대상 시트를 읽어 실제 파일/폴더를 삭제합니다.

실행 방법:
    python MEG_ChatBot/data/MEG_STANDARD/preprocess/src/delete_apply.py

동작 규칙:
    - [삭제대상_비엑셀파일] / [삭제대상_키워드] 시트에서
      '★ 삭제 여부' 칸이 Y (또는 y) 인 항목만 삭제
    - N 으로 변경한 항목은 건드리지 않음
    - 폴더 삭제 시 하위 내용 포함 전체 삭제 (shutil.rmtree)
    - dry-run 먼저 출력 후 실행 여부 확인

주의:
    삭제는 휴지통이 아닌 완전 삭제입니다.
    delete_apply.py 실행 전에 중요한 파일이 없는지 반드시 확인하세요.
"""

import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).resolve().parent
PREPROCESS_DIR = SCRIPT_DIR.parent
MEG_STD_DIR    = PREPROCESS_DIR.parent
RAW_DATA_ROOT  = MEG_STD_DIR / "raw_data"
ANALYSIS_FILE  = PREPROCESS_DIR / "result" / "name_analysis.xlsx"

# 삭제 대상 시트 이름
DELETE_SHEETS = ["삭제대상_비엑셀파일", "삭제대상_키워드"]

# 각 시트의 컬럼 구성 (1-based)
# 공통: 깊이(1) 구분(2) 이름(3) 이유(4) [키워드시트: 매칭키워드(5)] 상대경로(?) ★삭제여부(마지막)
# → 상대 경로와 ★ 삭제 여부 위치는 시트별로 다름, 마지막에서 역순으로 찾음


def load_delete_plan(xlsx_path: Path) -> list[dict]:
    """
    두 삭제 대상 시트에서 '★ 삭제 여부 = Y' 인 항목을 읽어 반환.
    반환: [{"path": Path, "name": str, "item_type": str, "reason": str, "sheet": str}]
    """
    wb   = load_workbook(xlsx_path, data_only=True)
    plan = []

    for sheet_name in DELETE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[경고] 시트 없음: {sheet_name} — 스킵")
            continue

        ws = wb[sheet_name]
        # 헤더 행(1행)에서 각 컬럼 위치 동적 탐지
        header_row  = [c.value for c in ws[1]]
        col_rel     = next((i for i, v in enumerate(header_row) if v and "상대 경로" in str(v)), None)
        col_yn      = next((i for i, v in enumerate(header_row) if v and "삭제 여부" in str(v)), None)
        col_type    = 1   # 구분 컬럼 (0-based index = 1)

        if col_rel is None or col_yn is None:
            print(f"[경고] {sheet_name}: 컬럼을 찾을 수 없음 — 스킵")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            yn_val   = str(row[col_yn]).strip().upper() if row[col_yn] else ""
            rel_path = str(row[col_rel]).strip()        if row[col_rel] else ""
            item_type = str(row[col_type]).strip()      if row[col_type] else ""
            reason   = str(row[3]).strip()              if row[3] else ""

            if yn_val != "Y":
                continue
            if not rel_path or rel_path in ("None", ""):
                continue

            target_path = RAW_DATA_ROOT / rel_path
            plan.append({
                "path":      target_path,
                "name":      target_path.name,
                "item_type": item_type,
                "reason":    reason,
                "sheet":     sheet_name,
                "rel_path":  rel_path,
            })

    # 중복 제거 (키워드+비엑셀 양쪽에 걸린 경우)
    seen = set()
    unique_plan = []
    for item in plan:
        key = str(item["path"])
        if key not in seen:
            seen.add(key)
            unique_plan.append(item)

    return unique_plan


def sort_by_depth_desc(plan: list[dict]) -> list[dict]:
    """하위 항목부터 삭제 (상위 폴더가 먼저 지워지면 하위 경로가 없어지는 문제 방지)."""
    return sorted(plan, key=lambda x: x["rel_path"].replace("\\", "/").count("/"), reverse=True)


def print_plan(plan: list[dict]):
    print(f"\n{'=' * 70}")
    print(f"  삭제 예정 항목: {len(plan)}개")
    print(f"{'=' * 70}")
    for item in plan:
        icon = "📁" if item["item_type"] == "폴더" else "📄"
        exists = item["path"].exists()
        status = "" if exists else "  ※ 이미 없음"
        print(f"  {icon} [{item['sheet']}] {item['reason']}")
        print(f"     {item['rel_path']}{status}")
    print()


def apply_deletes(plan: list[dict]) -> tuple[int, int, int]:
    success = skip = fail = 0

    for item in plan:
        p = item["path"]

        if not p.exists():
            print(f"[스킵-없음] {item['rel_path']}")
            skip += 1
            continue

        try:
            if p.is_dir():
                shutil.rmtree(p)
            else:
                p.unlink()
            print(f"[삭제완료] {item['rel_path']}")
            success += 1
        except Exception as e:
            print(f"[실패] {item['rel_path']}  —  {e}")
            fail += 1

    return success, skip, fail


def main():
    if not ANALYSIS_FILE.exists():
        print(f"❌ 분석 파일을 찾을 수 없습니다: {ANALYSIS_FILE}")
        sys.exit(1)
    if not RAW_DATA_ROOT.exists():
        print(f"❌ raw_data 폴더를 찾을 수 없습니다: {RAW_DATA_ROOT}")
        sys.exit(1)

    print(f"📄 엑셀 읽는 중: {ANALYSIS_FILE}")
    plan = load_delete_plan(ANALYSIS_FILE)

    if not plan:
        print("✅ 삭제할 항목이 없습니다. (모두 N 처리됐거나 시트가 비어 있음)")
        sys.exit(0)

    plan = sort_by_depth_desc(plan)
    print_plan(plan)

    # 경고 메시지
    print("⚠️  주의: 삭제는 휴지통이 아닌 완전 삭제입니다.")
    print("      폴더 삭제 시 하위 파일 전체가 함께 삭제됩니다.\n")

    answer = input("위 항목들을 완전 삭제할까요? (yes / no) > ").strip().lower()
    if answer not in ("yes", "y"):
        print("취소됐습니다. 변경 사항 없음.")
        sys.exit(0)

    print(f"\n{'─' * 70}")
    print("삭제 시작...\n")
    success, skip, fail = apply_deletes(plan)

    print(f"\n{'=' * 70}")
    print(f"  완료: {success}개 | 스킵: {skip}개 | 실패: {fail}개")
    print(f"{'=' * 70}")
    if fail > 0:
        print("⚠️  실패 항목이 있습니다. 위 로그를 확인하세요.")


if __name__ == "__main__":
    main()
