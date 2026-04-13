"""
analyze_names.py v2
"""
import re, sys
from pathlib import Path
from dataclasses import dataclass, field
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR     = Path(__file__).resolve().parent
PREPROCESS_DIR = SCRIPT_DIR.parent
MEG_STD_DIR    = PREPROCESS_DIR.parent
RAW_DATA_ROOT  = MEG_STD_DIR / "raw_data"
RESULT_DIR     = PREPROCESS_DIR / "result"

ALLOWED_EXTENSIONS = {".xlsx"}
DELETE_KEYWORDS    = ["개선", "old", "OLD", "삭제"]

PROTECT_PATTERNS = [
    r'2\.5\s*[Dd]', r'(?<!\d)3\s*[Dd](?!\w)', r'(?<!\d)2\s*[Dd](?!\w)',
    r'\d+[Dd][Oo][Tt]\d+', r'\d+\.\d+[Dd]',
    r'[Vv]\d+[\._]\d+', r'[Rr][Ee][Vv]\d+',
    r'IP\d+', r'USB\d+', r'WiFi\d', r'UWB', r'NFC', r'BT\d',
    r'\d+\s*[Mm][Mm]', r'\d+[Tt](?!\w)', r'\d+[Xx]\d+',
    r'\d+\.\d+["\'인치]', r'\d+["\'인치]', r'[A-Z]{2,3}\d{2,}',
]
_PROTECT_RE = [re.compile(p) for p in PROTECT_PATTERNS]

PREFIX_RULES = [
    ("숫자-숫자-  (예: 1-1- )",  re.compile(r'^(\d+[\-\.]\d+[\-\._]\s*)')),
    ("숫자-       (예: 1- )",     re.compile(r'^(\d+[\-]\s*)')),
    ("숫자.       (예: 1. )",     re.compile(r'^(\d+\.\s*)')),
    ("숫자_       (예: 01_ )",    re.compile(r'^(\d+_\s*)')),
    ("숫자공백    (예: 01 설계)", re.compile(r'^(\d+\s+)')),
    ("(숫자)      (예: (1) )",    re.compile(r'^(\(\d+\)\s*)')),
    ("[숫자]      (예: [01] )",   re.compile(r'^(\[\d+\]\s*)')),
]
SUFFIX_RULES = [
    ("_숫자 (끝부분, 예: _01)", re.compile(r'(_\d+)$')),
    ("-숫자 (끝부분, 예: -01)", re.compile(r'(-\d+)$')),
]
MID_RULES = [
    ("단어 사이 _ → 공백",    re.compile(r'(?<=\S)_(?=\S)'),  True,  " "),
    ("괄호 (예: 설계(검토))", re.compile(r'[\(\)\[\]]'),       False, ""),
    ("연속 언더스코어/하이픈", re.compile(r'[-_]{2,}'),         False, ""),
]

def _apply_mid_rules(working, rec):
    for rule in MID_RULES:
        desc, pat = rule[0], rule[1]
        auto = rule[2] if len(rule) > 2 else False
        repl = rule[3] if len(rule) > 3 else ""
        if pat.search(working):
            if auto:
                working = pat.sub(repl, working)
                rec.detected_patterns.append(desc)
            else:
                rec.detected_patterns.append(f"[중간/정보확인] {desc}")
                rec.needs_review = True
    return working

def has_protected_number(name):
    matched = []
    for pat in _PROTECT_RE:
        m = pat.search(name)
        if m:
            matched.append(m.group())
    return bool(matched), matched

@dataclass
class NameRecord:
    item_type: str; depth: int; original: str; stem: str
    extension: str; relative_path: str
    detected_patterns: list = field(default_factory=list)
    matched_prefix: str = ""; suggested: str = ""
    protected: bool = False; protected_keywords: list = field(default_factory=list)
    needs_review: bool = False

@dataclass
class DeleteRecord:
    item_type: str; depth: int; name: str; extension: str
    relative_path: str; reason: str; matched_keyword: str

def analyze_name(name, item_type, depth, rel_path):
    path_obj  = Path(name)
    stem      = path_obj.stem if item_type == "파일" else name
    extension = path_obj.suffix if item_type == "파일" else ""
    rec = NameRecord(item_type=item_type, depth=depth, original=name,
                     stem=stem, extension=extension, relative_path=rel_path)
    is_protected, prot_kw = has_protected_number(stem)
    rec.protected, rec.protected_keywords = is_protected, prot_kw
    working = stem

    for desc, pat in PREFIX_RULES:
        m = pat.match(working)
        if not m: continue
        if is_protected:
            after = pat.sub('', working).strip()
            if not any(p.search(after) for p in _PROTECT_RE):
                rec.detected_patterns.append(f"[보호됨] {desc}")
                rec.needs_review = True
                continue
        rec.detected_patterns.append(desc)
        rec.matched_prefix = m.group(1)
        working = pat.sub('', working).strip()
        break

    for desc, pat in SUFFIX_RULES:
        m = pat.search(working)
        if not m: continue
        if any(p.search(m.group(1)) for p in _PROTECT_RE):
            rec.detected_patterns.append(f"[보호확인필요] {desc}")
            rec.needs_review = True
        else:
            rec.detected_patterns.append(desc)
            working = pat.sub('', working).strip()

    working = _apply_mid_rules(working, rec)
    if is_protected and rec.matched_prefix:
        rec.needs_review = True
    working = working.strip()
    rec.suggested = working + extension
    if not rec.detected_patterns:
        rec.detected_patterns = ["패턴 없음"]
    return rec

def check_delete_targets(path, name, item_type, depth, rel_path):
    ext = path.suffix.lower() if item_type == "파일" else ""
    non_excel_rec = keyword_rec = None
    if item_type == "파일" and ext not in ALLOWED_EXTENSIONS:
        non_excel_rec = DeleteRecord(
            item_type=item_type, depth=depth, name=name, extension=ext,
            relative_path=rel_path,
            reason=f"비허용 확장자 ({ext if ext else '확장자 없음'})",
            matched_keyword="")
    for kw in DELETE_KEYWORDS:
        if kw in name:
            keyword_rec = DeleteRecord(
                item_type=item_type, depth=depth, name=name, extension=ext,
                relative_path=rel_path, reason="키워드 포함", matched_keyword=kw)
            break
    return non_excel_rec, keyword_rec

def scan_and_analyze(root):
    name_records, non_excel_list, keyword_list = [], [], []
    def _walk(path, depth):
        try: children = sorted(path.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
        except PermissionError: return
        for child in children:
            rel       = str(child.relative_to(root))
            item_type = "파일" if child.is_file() else "폴더"
            name_records.append(analyze_name(child.name, item_type, depth, rel))
            ne, kw = check_delete_targets(child, child.name, item_type, depth, rel)
            if ne: non_excel_list.append(ne)
            if kw: keyword_list.append(kw)
            if child.is_dir(): _walk(child, depth + 1)
    _walk(root, 1)
    return name_records, non_excel_list, keyword_list

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
C_HEADER="1F4E79"; C_FOLDER="D9E1F2"; C_FILE="FFFFFF"; C_WARN="FFF2CC"
C_PROTECTED="E2EFDA"; C_NO_PATTERN="F2F2F2"; C_INPUT="FFFBE6"; C_DELETE="FCE4D6"

def hfill(c): return PatternFill("solid", start_color=c)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = hfill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width: ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value, bg=C_FILE, bold=False, color="000000", align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.fill = hfill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    return c

def write_delete_sheet(ws, records, is_keyword=False):
    hdrs = [("깊이",5),("구분",7),("이름",45),("이유",30)]
    if is_keyword: hdrs.append(("매칭 키워드",16))
    hdrs += [("상대 경로",60),("★ 삭제 여부 (Y=삭제 / N=유지)",32)]
    for col,(h,w) in enumerate(hdrs,1): hdr_cell(ws,1,col,h,w)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for r_idx, rec in enumerate(records, 2):
        is_f = rec.item_type == "폴더"
        c = 1
        data_cell(ws,r_idx,c,rec.depth,     bg=C_DELETE,align="center");         c+=1
        data_cell(ws,r_idx,c,rec.item_type, bg=C_DELETE,align="center",
                  bold=is_f,color="1F4E79" if is_f else "000000");                c+=1
        data_cell(ws,r_idx,c,rec.name,      bg=C_DELETE,
                  bold=is_f,color="1F4E79" if is_f else "000000");                c+=1
        data_cell(ws,r_idx,c,rec.reason,    bg=C_DELETE,color="C00000");          c+=1
        if is_keyword:
            data_cell(ws,r_idx,c,rec.matched_keyword,bg=C_DELETE,bold=True,color="C00000"); c+=1
        data_cell(ws,r_idx,c,rec.relative_path,bg=C_DELETE,color="595959");       c+=1
        yn = ws.cell(row=r_idx, column=c, value="Y")
        yn.font = Font(name="Arial", bold=True, size=11, color="C00000")
        yn.fill = hfill(C_DELETE)
        yn.alignment = Alignment(horizontal="center", vertical="center")
        yn.border = BORDER
        ws.row_dimensions[r_idx].height = 18

    note_row = len(records) + 3
    n = ws.cell(row=note_row, column=1,
                value="※ 유지할 항목은 '★ 삭제 여부' 칸을 N 으로 변경하세요. 저장 후 delete_apply.py 실행.")
    n.font = Font(name="Arial", italic=True, size=9, color="595959")

def save_excel(name_records, non_excel_list, keyword_list, output_path):
    wb = Workbook()

    # 시트1: 분석 결과
    ws = wb.active; ws.title = "분석 결과"; ws.freeze_panes = "A2"
    HDRS = [("깊이",6),("구분",7),("원본 이름",42),("감지된 패턴",38),
            ("제안 정제명",42),("★ 확정 이름 (직접 입력)",42),("보호 키워드",28),("상대 경로",55)]
    for col,(h,w) in enumerate(HDRS,1): hdr_cell(ws,1,col,h,w)
    ws.row_dimensions[1].height = 30
    pat_ctr = Counter(); prot_recs = []

    for r_idx, rec in enumerate(name_records, 2):
        ps = " | ".join(rec.detected_patterns)
        if "패턴 없음" in ps:           bg = C_NO_PATTERN
        elif rec.needs_review or rec.protected: bg = C_WARN
        elif rec.item_type == "폴더":   bg = C_FOLDER
        else:                           bg = C_FILE
        changed = rec.suggested != rec.original
        sd = rec.suggested if changed else "- (변경 없음)"
        is_f = rec.item_type == "폴더"
        data_cell(ws,r_idx,1,rec.depth,    bg=bg,align="center")
        data_cell(ws,r_idx,2,rec.item_type,bg=bg,align="center",bold=is_f,color="1F4E79" if is_f else "000000")
        data_cell(ws,r_idx,3,rec.original, bg=bg,bold=is_f,color="1F4E79" if is_f else "000000")
        data_cell(ws,r_idx,4,ps,           bg=bg,color="C00000" if (rec.needs_review or (changed and "패턴 없음" not in ps)) else "595959")
        data_cell(ws,r_idx,5,sd,           bg=C_PROTECTED if rec.protected else bg,bold=changed,
                  color="375623" if rec.protected else ("000000" if not changed else "244061"))
        data_cell(ws,r_idx,6,"",           bg=C_INPUT)
        data_cell(ws,r_idx,7,", ".join(rec.protected_keywords) if rec.protected_keywords else "",bg=bg,color="375623")
        data_cell(ws,r_idx,8,rec.relative_path,bg=bg,color="595959")
        ws.row_dimensions[r_idx].height = 18
        for p in rec.detected_patterns: pat_ctr[p] += 1
        if rec.protected: prot_recs.append(rec)

    # 시트2: 범례
    ws2 = wb.create_sheet("범례 및 가이드")
    ws2.column_dimensions["A"].width=22; ws2.column_dimensions["B"].width=55; ws2.column_dimensions["C"].width=38
    guide = [
        ("색상 범례","",""),
        ("노란 행","검토 필요 — 보호 키워드 포함 or 복합 패턴","직접 판단 후 확정 이름 칸에 입력"),
        ("연초록 셀","제안 정제명 칸 — 보호 키워드가 있어 일부만 제거됨","보호된 숫자가 맞는지 확인"),
        ("회색 행","패턴 없음 — 이름이 깨끗함","별도 조치 불필요"),
        ("연파랑 행","폴더 (패턴 감지됨)","제안명 확인 후 확정"),
        ("흰색 행","파일 (패턴 감지됨)","제안명 확인 후 확정"),
        ("연노랑 셀","★ 확정 이름 입력칸","최종 사용할 이름을 직접 입력"),
        ("연주황 행","삭제 대상 시트 — 비엑셀 파일 or 키워드 포함","유지하려면 ★ 삭제 여부를 N 으로"),
        ("","",""),
        ("패턴 설명","",""),
        ("숫자-","1-, 01- 등 앞에 붙은 숫자+하이픈","ex) 1-설계  →  설계"),
        ("숫자.","1. 등 앞에 붙은 숫자+점","ex) 1.낙하  →  낙하"),
        ("숫자_","01_ 등 앞에 붙은 숫자+언더스코어","ex) 01_낙하  →  낙하"),
        ("숫자-숫자-","1-1-, 2-3- 등 복합 접두","ex) 1-1-설계  →  설계"),
        ("단어 사이 _","단어 사이의 _ 를 공백으로 자동 치환","ex) 단차_gap  →  단차 gap"),
        ("[보호됨]","삭제 시 의미 손실 우려 — 수동 판단 필요","ex) IP67, 2.5D, Rev3"),
        ("[중간/정보확인]","이름 중간의 괄호·특수문자","ex) 설계(최종)"),
        ("","",""),
        ("삭제 대상 설명","",""),
        ("비엑셀 파일",".xlsx 이외의 모든 파일","ex) .pdf .png .pptx 등"),
        ("키워드 포함","개선 / old / OLD / 삭제 가 이름에 포함된 항목","폴더·파일 모두 해당"),
        ("","",""),
        ("작업 순서","",""),
        ("1단계","[분석 결과] 노란 행 검토 → 확정 이름 입력",""),
        ("2단계","[삭제대상_비엑셀파일] 유지할 항목 N 으로 변경",""),
        ("3단계","[삭제대상_키워드] 유지할 항목 N 으로 변경",""),
        ("4단계","저장 후 rename_apply.py 실행","이름 변경"),
        ("5단계","저장 후 delete_apply.py 실행","삭제 처리"),
    ]
    cmap = {"색상 범례":C_HEADER,"패턴 설명":C_HEADER,"삭제 대상 설명":C_HEADER,"작업 순서":C_HEADER,
            "노란 행":C_WARN,"연초록 셀":C_PROTECTED,"회색 행":C_NO_PATTERN,
            "연파랑 행":C_FOLDER,"흰색 행":C_FILE,"연노랑 셀":C_INPUT,"연주황 행":C_DELETE}
    for ri,(a,b,c) in enumerate(guide,1):
        bg = cmap.get(a,"FFFFFF"); is_h = bg==C_HEADER
        for ci,v in enumerate([a,b,c],1):
            cell=ws2.cell(row=ri,column=ci,value=v)
            cell.font=Font(name="Arial",bold=is_h or (ci==1 and a),color="FFFFFF" if is_h else "000000",size=10)
            cell.fill=hfill(bg); cell.alignment=Alignment(vertical="center",wrap_text=True); cell.border=BORDER
        ws2.row_dimensions[ri].height=20

    # 시트3: 패턴 통계
    ws3 = wb.create_sheet("패턴 통계")
    hdr_cell(ws3,1,1,"감지된 패턴",40); hdr_cell(ws3,1,2,"출현 횟수",14); hdr_cell(ws3,1,3,"비율(%)",12)
    ws3.row_dimensions[1].height=22
    total=sum(pat_ctr.values())
    for ri,(p,cnt) in enumerate(pat_ctr.most_common(),2):
        data_cell(ws3,ri,1,p); data_cell(ws3,ri,2,cnt,align="center")
        data_cell(ws3,ri,3,f"{cnt/total*100:.1f}%" if total else "-",align="center")
        ws3.row_dimensions[ri].height=18

    # 시트4: 보호 키워드
    ws4 = wb.create_sheet("보호 키워드 목록")
    hdr_cell(ws4,1,1,"원본 이름",40); hdr_cell(ws4,1,2,"보호된 키워드",30)
    hdr_cell(ws4,1,3,"제안 정제명",40); hdr_cell(ws4,1,4,"상대 경로",55)
    ws4.row_dimensions[1].height=22
    for ri,rec in enumerate(prot_recs,2):
        data_cell(ws4,ri,1,rec.original,bg=C_WARN,bold=True)
        data_cell(ws4,ri,2,", ".join(rec.protected_keywords),bg=C_PROTECTED,color="375623")
        data_cell(ws4,ri,3,rec.suggested,bg=C_WARN)
        data_cell(ws4,ri,4,rec.relative_path,color="595959")
        ws4.row_dimensions[ri].height=18

    # 시트5: 삭제대상_비엑셀파일
    ws5 = wb.create_sheet("삭제대상_비엑셀파일")
    if non_excel_list: write_delete_sheet(ws5, non_excel_list, is_keyword=False)
    else: ws5.cell(row=1,column=1,value="비엑셀 파일 없음 ✅").font=Font(name="Arial",bold=True,color="375623")

    # 시트6: 삭제대상_키워드
    ws6 = wb.create_sheet("삭제대상_키워드")
    if keyword_list: write_delete_sheet(ws6, keyword_list, is_keyword=True)
    else: ws6.cell(row=1,column=1,value="키워드 포함 항목 없음 ✅").font=Font(name="Arial",bold=True,color="375623")

    wb.save(output_path)
    print(f"✅ 분석 결과 저장: {output_path}")

def main():
    if not RAW_DATA_ROOT.exists():
        print(f"❌ raw_data 폴더를 찾을 수 없습니다: {RAW_DATA_ROOT}"); sys.exit(1)
    print(f"📂 탐색 및 분석 시작: {RAW_DATA_ROOT}")
    name_records, non_excel_list, keyword_list = scan_and_analyze(RAW_DATA_ROOT)
    total=len(name_records)
    print(f"\n── 분석 요약 ─────────────────────────────────────────────")
    print(f"  전체 항목              : {total}개")
    print(f"  이름 패턴 감지         : {sum(1 for r in name_records if '패턴 없음' not in r.detected_patterns)}개")
    print(f"  수동 검토 필요 (노란행): {sum(1 for r in name_records if r.needs_review)}개")
    print(f"  보호 키워드 포함       : {sum(1 for r in name_records if r.protected)}개")
    print(f"  비엑셀 파일 (삭제대상) : {len(non_excel_list)}개")
    print(f"  키워드 포함 (삭제대상) : {len(keyword_list)}개")
    print(f"──────────────────────────────────────────────────────────")
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = RESULT_DIR / "name_analysis.xlsx"
    save_excel(name_records, non_excel_list, keyword_list, output_path)
    print(f"\n다음 단계:")
    print(f"  1. {output_path} 열기")
    print(f"  2. [분석 결과] 노란 행 검토")
    print(f"  3. [삭제대상_비엑셀파일] / [삭제대상_키워드] 확인 → 유지 항목 N 으로")
    print(f"  4. 저장 후 rename_apply.py → delete_apply.py 순으로 실행")

if __name__ == "__main__":
    main()
