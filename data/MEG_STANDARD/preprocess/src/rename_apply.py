"""
rename_apply.py
---------------
name_analysis.xlsx 를 읽어 실제 폴더/파일 이름을 변경합니다.

실행 방법:
    python MEG_ChatBot/data/MEG_STANDARD/preprocess/src/rename_apply.py

동작 규칙:
    1. ★ 확정 이름 칸이 입력돼 있으면  → 확정 이름으로 변경
    2. 확정 이름 칸이 비어 있고 제안 정제명이 "- (변경 없음)" 이 아니면  → 제안 정제명으로 변경
    3. 제안 정제명이 "- (변경 없음)" 이면  → 건드리지 않음

실행 순서:
    1. dry-run 모드로 먼저 실행 → 변경 예정 목록 출력 (실제 변경 없음)
    2. 내용 확인 후 실제 실행 여부 입력
"""

import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).resolve().parent
PREPROCESS_DIR = SCRIPT_DIR.parent
MEG_STD_DIR    = PREPROCESS_DIR.parent
RAW_DATA_ROOT  = MEG_STD_DIR / "raw_data"
ANALYSIS_FILE  = PREPROCESS_DIR / "result" / "name_analysis.xlsx"

# ── 엑셀 컬럼 인덱스 (1-based) ────────────────────────────────────────────────
COL_ORIGINAL  = 3   # 원본 이름
COL_SUGGESTED = 5   # 제안 정제명
COL_CONFIRMED = 6   # ★ 확정 이름 (사용자 입력)
COL_REL_PATH  = 8   # 상대 경로


# ── 엑셀 읽기 ──────────────────────────────────────────────────────────────────
def load_rename_plan(xlsx_path: Path) -> list[dict]:
    """
    엑셀에서 변경 계획을 읽어 반환.
    반환 형태: [{"original": Path, "new_name": str}, ...]
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["분석 결과"]

    plan = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        original_name = row[COL_ORIGINAL - 1]
        suggested     = row[COL_SUGGESTED - 1]
        confirmed     = row[COL_CONFIRMED - 1]
        rel_path      = row[COL_REL_PATH  - 1]

        if not original_name or not rel_path:
            continue

        # 확정 이름 우선, 없으면 제안명 사용
        confirmed_str = str(confirmed).strip() if confirmed else ""
        suggested_str = str(suggested).strip() if suggested else ""

        if confirmed_str:
            new_name = confirmed_str
        elif suggested_str and suggested_str != "- (변경 없음)":
            new_name = suggested_str
        else:
            continue  # 변경 없음

        # 원본과 동일하면 스킵
        if new_name == str(original_name).strip():
            continue

        original_path = RAW_DATA_ROOT / rel_path
        plan.append({
            "original_path": original_path,
            "original_name": str(original_name).strip(),
            "new_name":      new_name,
            "rel_path":      rel_path,
        })

    return plan


# ── 충돌 검사 ──────────────────────────────────────────────────────────────────
def check_conflicts(plan: list[dict]) -> list[dict]:
    """변경 후 이름이 이미 존재하는 경우를 찾아 반환."""
    conflicts = []
    for item in plan:
        orig = item["original_path"]
        new_path = orig.parent / item["new_name"]
        if new_path.exists() and new_path != orig:
            conflicts.append({**item, "conflict_path": new_path})
    return conflicts


# ── 실행 순서 정렬 ─────────────────────────────────────────────────────────────
def sort_by_depth_desc(plan: list[dict]) -> list[dict]:
    """
    깊은 경로부터 변경해야 상위 폴더 이름 변경 시 하위 경로가 깨지지 않음.
    rel_path 의 구분자(/ or \\) 수로 깊이 판단.
    """
    def depth(item):
        p = item["rel_path"].replace("\\", "/")
        return p.count("/")

    return sorted(plan, key=depth, reverse=True)


# ── dry-run 출력 ───────────────────────────────────────────────────────────────
def print_plan(plan: list[dict], conflicts: list[dict]):
    conflict_paths = {c["original_path"] for c in conflicts}

    print(f"\n{'=' * 70}")
    print(f"  변경 예정 항목: {len(plan)}개")
    if conflicts:
        print(f"  ⚠️  충돌 항목:    {len(conflicts)}개  (아래 ⚠️  표시)")
    print(f"{'=' * 70}")

    for item in plan:
        is_conflict = item["original_path"] in conflict_paths
        icon = "⚠️ " if is_conflict else "  "
        print(f"{icon}[{item['rel_path']}]")
        print(f"      {item['original_name']}")
        print(f"   →  {item['new_name']}")
        if is_conflict:
            print(f"      ※ 충돌: 변경 후 이름이 이미 존재합니다 — 스킵됩니다")
        print()


# ── 실제 변경 ──────────────────────────────────────────────────────────────────
def apply_renames(plan: list[dict], conflicts: list[dict]) -> tuple[int, int, int]:
    conflict_paths = {c["original_path"] for c in conflicts}
    success = skip = fail = 0

    for item in plan:
        orig = item["original_path"]
        new_path = orig.parent / item["new_name"]

        # 충돌 항목 스킵
        if orig in conflict_paths:
            print(f"[스킵-충돌] {item['original_name']}")
            skip += 1
            continue

        # 원본 경로가 없는 경우 (상위 폴더가 이미 변경된 경우 등)
        if not orig.exists():
            print(f"[스킵-없음] {orig}  (이미 이동됐거나 경로 불일치)")
            skip += 1
            continue

        try:
            orig.rename(new_path)
            print(f"[완료] {item['original_name']}  →  {item['new_name']}")
            success += 1
        except Exception as e:
            print(f"[실패] {item['original_name']}  —  {e}")
            fail += 1

    return success, skip, fail


# ── 메인 ───────────────────────────────────────────────────────────────────────
def main():
    # 파일 존재 확인
    if not ANALYSIS_FILE.exists():
        print(f"❌ 분석 파일을 찾을 수 없습니다: {ANALYSIS_FILE}")
        sys.exit(1)
    if not RAW_DATA_ROOT.exists():
        print(f"❌ raw_data 폴더를 찾을 수 없습니다: {RAW_DATA_ROOT}")
        sys.exit(1)

    print(f"📄 엑셀 읽는 중: {ANALYSIS_FILE}")
    plan = load_rename_plan(ANALYSIS_FILE)

    if not plan:
        print("✅ 변경할 항목이 없습니다.")
        sys.exit(0)

    # 깊은 경로부터 처리 (하위 → 상위 순서)
    plan = sort_by_depth_desc(plan)

    # 충돌 검사
    conflicts = check_conflicts(plan)

    # dry-run 출력
    print_plan(plan, conflicts)

    # 충돌 경고
    if conflicts:
        print(f"⚠️  충돌 항목 {len(conflicts)}개는 자동으로 스킵됩니다.")
        print(f"   엑셀에서 해당 항목의 확정 이름을 수정 후 재실행하세요.\n")

    # 실행 여부 확인
    answer = input("위 내용으로 실제 변경을 진행할까요? (yes / no) > ").strip().lower()
    if answer not in ("yes", "y"):
        print("취소됐습니다. 변경 사항 없음.")
        sys.exit(0)

    # 실제 변경
    print(f"\n{'─' * 70}")
    print("변경 시작...\n")
    success, skip, fail = apply_renames(plan, conflicts)

    print(f"\n{'=' * 70}")
    print(f"  완료: {success}개 | 스킵: {skip}개 | 실패: {fail}개")
    print(f"{'=' * 70}")

    if fail > 0:
        print("⚠️  실패 항목이 있습니다. 위 로그를 확인하세요.")


if __name__ == "__main__":
    main()
