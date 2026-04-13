"""
scan_raw_data.py
----------------
MEG_ChatBot/data/MEG_STANDARD/raw_data/ 하위의 모든 폴더·파일을
재귀적으로 탐색하여 엑셀로 저장합니다.

저장 위치: MEG_ChatBot/data/MEG_STANDARD/preprocess/result/raw_data_structure.xlsx

실행 방법:
    python MEG_ChatBot/data/MEG_STANDARD/preprocess/src/scan_raw_data.py
또는 이 파일을 src 폴더에 놓고:
    python scan_raw_data.py
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
# 이 스크립트 위치: MEG_ChatBot/data/MEG_STANDARD/preprocess/src/
SCRIPT_DIR = Path(__file__).resolve().parent          # .../preprocess/src/
PREPROCESS_DIR = SCRIPT_DIR.parent                    # .../preprocess/
MEG_STANDARD_DIR = PREPROCESS_DIR.parent              # .../MEG_STANDARD/

RAW_DATA_ROOT = MEG_STANDARD_DIR / "raw_data"
RESULT_DIR    = PREPROCESS_DIR / "result"

# ── 폴더 탐색 ──────────────────────────────────────────────────────────────────
def scan_directory(root: Path):
    """
    root 하위를 재귀 탐색하여 폴더 정보 리스트를 반환합니다.
    반환 형태: [
        {
            "depth": int,              # 깊이 (root = 0)
            "type": "folder" | "file",
            "name": str,               # 파일/폴더명
            "relative_path": str,      # root 기준 상대 경로
            "absolute_path": str,
            "parent_folder": str,      # 바로 위 부모 폴더명
            "item_count": int | None,  # 폴더인 경우 직계 자식 수
        }
    ]
    """
    records = []

    def _walk(path: Path, depth: int, parent_name: str):
        try:
            children = sorted(path.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
        except PermissionError:
            return

        for child in children:
            rel = child.relative_to(root)
            record = {
                "depth": depth,
                "type": "folder" if child.is_dir() else "file",
                "name": child.name,
                "relative_path": str(rel),
                "absolute_path": str(child),
                "parent_folder": parent_name,
                "item_count": None,
            }
            if child.is_dir():
                try:
                    record["item_count"] = sum(1 for _ in child.iterdir())
                except PermissionError:
                    record["item_count"] = -1
                records.append(record)
                _walk(child, depth + 1, child.name)
            else:
                records.append(record)

    _walk(root, 1, root.name)
    return records


# ── 엑셀 저장 ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # 진한 파랑
FOLDER_FILL   = PatternFill("solid", start_color="D9E1F2")   # 연한 파랑
FILE_FILL     = PatternFill("solid", start_color="FFFFFF")   # 흰색
DEPTH_COLORS  = [
    "BDD7EE", "9DC3E6", "2E75B6", "1F4E79",                 # depth 1~4
]
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["깊이(Depth)", "구분", "이름", "상위 폴더", "상대 경로", "직계 항목 수"]
COL_WIDTHS = [14, 10, 45, 30, 65, 14]

def depth_fill(depth: int) -> PatternFill:
    color = DEPTH_COLORS[min(depth - 1, len(DEPTH_COLORS) - 1)]
    return PatternFill("solid", start_color=color)

def save_to_excel(records: list, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "폴더 구조"

    # ── 헤더 ──
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    # ── 데이터 ──
    for row_idx, rec in enumerate(records, start=2):
        is_folder = rec["type"] == "folder"
        values = [
            rec["depth"],
            "📁 폴더" if is_folder else "📄 파일",
            ("    " * (rec["depth"] - 1)) + rec["name"],   # 들여쓰기로 계층 표현
            rec["parent_folder"],
            rec["relative_path"],
            rec["item_count"] if is_folder else "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(
                name="Arial",
                bold=is_folder,
                size=10,
                color="1F4E79" if is_folder else "000000",
            )
            cell.fill = depth_fill(rec["depth"]) if is_folder else FILE_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = BORDER

        ws.row_dimensions[row_idx].height = 18

    # ── 요약 시트 ──
    ws2 = wb.create_sheet("요약")
    folder_count = sum(1 for r in records if r["type"] == "folder")
    file_count   = sum(1 for r in records if r["type"] == "file")
    max_depth    = max((r["depth"] for r in records), default=0)

    summary_data = [
        ("항목", "값"),
        ("탐색 루트", str(RAW_DATA_ROOT)),
        ("총 폴더 수", folder_count),
        ("총 파일 수", file_count),
        ("최대 깊이",  max_depth),
    ]
    for r_idx, (k, v) in enumerate(summary_data, start=1):
        for c_idx, val in enumerate([k, v], start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", bold=(r_idx == 1), size=10,
                             color="FFFFFF" if r_idx == 1 else "000000")
            cell.fill = HEADER_FILL if r_idx == 1 else PatternFill("solid", start_color="EBF3FB")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
        ws2.row_dimensions[r_idx].height = 20

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 70

    wb.save(output_path)
    print(f"✅ 엑셀 저장 완료: {output_path}")


# ── 메인 ───────────────────────────────────────────────────────────────────────
def main():
    if not RAW_DATA_ROOT.exists():
        print(f"❌ raw_data 폴더를 찾을 수 없습니다: {RAW_DATA_ROOT}")
        sys.exit(1)

    print(f"📂 탐색 시작: {RAW_DATA_ROOT}")
    records = scan_directory(RAW_DATA_ROOT)

    folder_count = sum(1 for r in records if r["type"] == "folder")
    file_count   = sum(1 for r in records if r["type"] == "file")
    print(f"   → 폴더 {folder_count}개 / 파일 {file_count}개 발견")

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = RESULT_DIR / "raw_data_structure.xlsx"

    save_to_excel(records, output_path)

    # 간단한 트리 미리보기 (터미널 출력)
    print("\n── 폴더 트리 미리보기 ──────────────────────────────")
    for rec in records:
        if rec["type"] == "folder":
            indent = "  " * (rec["depth"] - 1)
            count_str = f"  [{rec['item_count']}개]" if rec["item_count"] is not None else ""
            print(f"{indent}📁 {rec['name']}{count_str}")
        else:
            indent = "  " * (rec["depth"] - 1)
            print(f"{indent}  └─ 📄 {rec['name']}")


if __name__ == "__main__":
    main()
